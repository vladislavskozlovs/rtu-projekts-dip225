#Definējam bibliotēkas
from openpyxl import load_workbook

#Atveram dokumentus
ievadfails = load_workbook('skoleni.xlsx')
ad_forma = load_workbook('ad.xlsx')
ms_forma = load_workbook('ms.xlsx')
#Izvēlamies atvērtās izklājlapas dokumenta lapu
datu_lapa = ievadfails.active
ad_lapa = ad_forma.active
ms_lapa = ms_forma.active

#Nosakam lapā kopīgo kolonu skaitu
#Ievērojam, ka Python skaitīšana sākās no 0
max_kolonu = datu_lapa.max_row + int(1)
max_kolonu_ad = ad_lapa.max_row + int(1)
max_kolonu_ms = ms_lapa.max_row + int(1)

#Definēju kolonu virsraktus izklājlapā
datu_lapa['E1'].value = "Vārds un Cits vārds"
datu_lapa['F1'].value = "Vārdi un uzvārdi"
datu_lapa['G1'].value = "Sistēmā uzrādāmis pilnais lietotājvārds"
datu_lapa['H1'].value = "Lietotājvārds"
datu_lapa['I1'].value = "Lietotāja e-pasts"
datu_lapa['J1'].value = "OU"

#Apvieno vārdus vienā laukā
for kolona in range(2, max_kolonu):
    pirmais_vards = datu_lapa['A' + str(kolona)].value
    otrais_vards = datu_lapa['B' + str(kolona)].value
    #Ja nav 2. vārda tad, neviedojās atstarpe starp 1. un 2. vārdu.
    #Saglabājas kolonā E izklājlapā
    if otrais_vards != None:
        datu_lapa['E' + str(kolona)].value = str(pirmais_vards + " " + otrais_vards)
    else:
        datu_lapa['E' + str(kolona)].value = pirmais_vards

#Apvieno saliktos vārdus un uzvārdu
#Saglabā datus izklājlapā F kolonā
for kolona in range(2, max_kolonu):
    vards = datu_lapa['E' + str(kolona)].value
    uzvards= datu_lapa['C' + str(kolona)].value
    datu_lapa['F' + str(kolona)].value= str(vards + " " + uzvards)

#Vārdi un uzvārdi sākās ar leilo burtu
#Likvidēti vārdos gaumzīmes un mīkstinājuma zīmes, atstarpes vietā ir punksts un visi burti ir maziņi
#No lietotājvārdiem izveidoti e-pasta adreses
for kolona in range(2, max_kolonu):
    #Programma G kolonai
    skolens_lielie_burti = datu_lapa['F' + str(kolona)].value
    datu_lapa['G' + str(kolona)].value = skolens_lielie_burti.title()
    sistemas_lietotajvards = datu_lapa['G' + str(kolona)].value
    #Programma H kolonai
    datu_lapa['H' + str(kolona)].value = skolens_lielie_burti.replace("Ā", "A").replace("Č", "C").replace("Ē", "E").replace("Ī", "I").replace("Ģ", "G").replace("Ķ", "K").replace("Ļ", "L").replace("Ņ", "N").replace("Š", "S").replace("Ū", "U").replace("Ž", "Z").replace(" ", ".").lower()
    skolens_ar_mazo_burtu = datu_lapa['H' + str(kolona)].value
    #Programma I kolonai
    datu_lapa['I' + str(kolona)].value = skolens_ar_mazo_burtu + "@edu.jtv.lv" 

#Definējam, kurā Organizāijas vienībā Aktīvajā Direktorijā atradīsies lietotājs, atbilstoši klasei un lietotāja ievadītajam mācību gadam.
macibu_gads = input("Ievadiet, lūdzu, kāds ir šobrīd mācību gads pēc šablona XXXX/XXXX: ")
for kolona in range(2, max_kolonu): 
    #Programma J kolonai
    klase= datu_lapa['D' + str(kolona)].value
    klase = klase.replace(".", "_").upper()
    datu_lapa['J' + str(kolona)].value = str("OU=" + macibu_gads + "_" + klase + ",OU=SKOLENI,OU=LIETOTAJI,OU=KONTI,DC=JTV,DC=LV")

#MS faila formējums
for kolona in range(2, max_kolonu):
        ms_lapa['A' + str(kolona)].value = datu_lapa['I' + str(kolona)].value
        ms_lapa['B' + str(kolona)].value = datu_lapa['E' + str(kolona)].value.title()
        ms_lapa['C' + str(kolona)].value = datu_lapa['C' + str(kolona)].value.title()
        ms_lapa['D' + str(kolona)].value = datu_lapa['G' + str(kolona)].value
        ms_lapa['E' + str(kolona)].value = "Izglītojamais"
        ms_lapa['L' + str(kolona)].value = "Meiju ceļš 9"
        ms_lapa['M' + str(kolona)].value = "Jelgava"

#AD faila formējums
for kolona in range(2, max_kolonu):
    ad_lapa['A' + str(kolona)].value = datu_lapa['E' + str(kolona)].value.title()
    ad_lapa['B' + str(kolona)].value = datu_lapa['C' + str(kolona)].value.title()
    ad_lapa['C' + str(kolona)].value = datu_lapa['G' + str(kolona)].value
    ad_lapa['D' + str(kolona)].value = datu_lapa['H' + str(kolona)].value
    ad_lapa['E' + str(kolona)].value = datu_lapa['I' + str(kolona)].value
    ad_lapa['H' + str(kolona)].value = "Jelgava"
    ad_lapa['L' + str(kolona)].value = "Izglītojamais"
    ad_lapa['P' + str(kolona)].value = datu_lapa['J' + str(kolona)].value
    ad_lapa['X' + str(kolona)].value = "Enable"

#Rezultāts saglabājās
ievadfails.save('result.xlsx')
ad_forma.save('ad_users.xlsx')
ievadfails.close()
ad_forma.close()